#routes/export.py

from flask import Blueprint, send_file, flash, redirect, url_for, request, session, jsonify, render_template, make_response
from io import BytesIO, StringIO
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime, date, timezone
from collections import OrderedDict
from babel.numbers import format_currency
from extensions import format_inr
from bson import ObjectId

import os, pandas as pd, subprocess, shutil, threading, time, schedule, csv, re, io, openpyxl

from utils import get_fields_for_type, normalize_cell, is_valid_date, is_future_date, get_master_fields, get_all_existing_types
from routes.main import safe_to_float, normalize_gst_keys
from models import assets_collection, asset_types_collection, import_previews_collection
from init_db import asset_type_fields


export_bp = Blueprint('export', __name__)

@export_bp.route('/debug-session')
def debug_session():
    return str(session)

# --- helper function for file check ---
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in {'xls', 'xlsx'}

# === 📥 1. EXPORT KEKA =======================================
@export_bp.route('/keka')
def export_keka():
    ids_param = request.args.get("ids")
    query = {}
    if ids_param:
        ids = [ObjectId(i) for i in ids_param.split(",") if i]
        query = {"_id": {"$in": ids}}

    assets = list(assets_collection.find(query))

    headers = [
        "Asset ID", "Asset Name", "Asset Description", "Asset Location", "Asset Category",
        "Asset Type", "Purchased On (dd-mmm-yyyy)", "Warranty Expires On (dd-mmm-yyyy)",
        "Asset Condition", "Asset Status", "Reason, if Not Available",
        "Employee Number, if Assigned", "Date of Asset Assignment (dd-mmm-yyyy)"
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "KEKA Export"

    bold_font = Font(bold=True, name="Calibri")
    wrap_align = Alignment(wrap_text=True, vertical="top", horizontal="left")

    ws.append(headers)
    for col in ws.iter_cols(min_row=1, max_row=1):
        for cell in col:
            cell.font = bold_font
            cell.alignment = wrap_align
            ws.column_dimensions[cell.column_letter].width = 25

    def fmt(val):
        return val if val else "-"

    def fmt_date(d):
        try:
            return datetime.strptime(str(d), "%d-%m-%Y").strftime("%d-%b-%Y")
        except Exception:
            return "-"

    for asset in assets:
        asset_name = fmt(str(asset.get("category") or "").strip())
        if asset_name.lower() == "desktop":
            cpu_tag = (
                str(asset.get("IT_tagC") or "").strip()
                or str(asset.get("accounts_tagC") or "").strip()
                or str(asset.get("endpoint_name") or "").strip()
                or "-"
            )

            monitor_or_mtr = (
                str(asset.get("IT_tagM") or "").strip()
                or str(asset.get("accounts_tagM") or "").strip()
                or str(asset.get("serial_no") or "").strip()
                or "-"
            )

            asset_id = f"{cpu_tag}, {monitor_or_mtr}"

        elif asset_name.lower() == "laptop":  # <-- ✅ Added Laptop logic
            # Asset ID: Prefer user_code, else serial_no
            asset_id = (
                str(asset.get("it_tag") or "").strip()
                or str(asset.get("accounts_tag") or "").strip()
                or str(asset.get("serial_no") or "").strip()
                or str(asset.get("endpoint_name") or "").strip()
                or "-"
            )

            # Asset Name: Combine manufacturer and model
            manufacturer = str(asset.get("system_manufacturer") or "").strip()
            model = str(asset.get("system_model") or "").strip()
            asset_name = ", ".join([m for m in [manufacturer, model] if m]) or asset_name


            # Asset Description: processor, RAM, OS, HDD, Free space, License
            desc_parts = [
                str(asset.get("processor") or "").strip(),
                str(asset.get("ram") or "").strip(),
                str(asset.get("os") or "").strip(),
                str(asset.get("hdd") or "").strip(),
                str(asset.get("license") or "").strip(),
            ]
            asset_desc = ", ".join([d for d in desc_parts if d]) or "-"

            # Asset Location: endpoint_name
            area = str(asset.get("area") or "").strip()
            state = str(asset.get("state") or "").strip()
            location = (f"{area} ({state})" if area and state else area or state or "-") #remove endpoint refernce here

        elif asset_name.lower() == "franchise inv":  # <-- Franchise Inventory
            asset_id = (
                str(asset.get("it_tag") or "").strip()
                or str(asset.get("accounts_tag") or "").strip()
                or str(asset.get("endpoint_name") or "").strip()
                or "-"
            )

        elif asset_name.lower() == "mobile":  # <-- Mobile logic
            imei1 = str(asset.get("imei1") or "").strip()
            imei2 = str(asset.get("imei2") or "").strip()
            if imei1:
                asset_id = imei1
            elif imei2:
                asset_id = imei2
            else:
                asset_id = "-"
        else:
            id_fields = [
                "IT_tagC", "accounts_tagC",
                "IT_tagM", "accounts_tagM",
                "it_tag", "accounts_tag",
                "serial_no", "endpoint_name"
            ]

            asset_id = next((str(asset.get(f) or "").strip() for f in id_fields if asset.get(f)), "")

            desc_parts = [
                str(asset.get("model") or "").strip(),
                str(asset.get("system_model") or "").strip(),
                str(asset.get("ram") or "").strip(),
                str(asset.get("storage") or "").strip()
            ]
            asset_desc = "  ".join([p for p in desc_parts if p]) or "-"

            area = str(asset.get("area") or "").strip()
            state = str(asset.get("state") or "").strip()
            location = f"{area} ({state})" if area and state else area or state or "-"

        
        asset_category = "IT assets"
        purchase_date = fmt_date(asset.get("purchase_date"))
        given_date = fmt_date(asset.get("given_date"))
        warranty_expires = "-"

        raw_status = str(asset.get("status") or "").strip().lower()
        valid_statuses = ["available(p)", "available(g)", "assigned(p)", "assigned(g)"]
        fallback_statuses = ["discard", "repair/faulty"]

        if raw_status in valid_statuses:
            suffix = raw_status.split("(")[1].replace(")", "") if "(" in raw_status else ""
            condition = "poor" if suffix == "p" else "good"
            status = raw_status
            remarks = "-"
        elif raw_status in fallback_statuses:
            condition = "-"
            status = "not available"
            remarks = raw_status
        else:
            condition = "-"
            status = fmt(raw_status)
            remarks = "-"

        username = str(asset.get("username") or "").strip() #employee name should come here
        user_code = str(asset.get("user_code") or "").strip() #employee code should come here
        if username and user_code:
            employee_number = f"{username} ({user_code})"
        else:
            employee_number = username or user_code or "-"

        assignment_date = given_date if given_date != "-" else "-"

        ws.append([
            fmt(asset_id),
            asset_name,
            asset_desc,
            location,
            asset_category,
            asset_name,
            purchase_date, 
            warranty_expires,
            condition,
            status,
            remarks,
            employee_number,
            assignment_date
        ])

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = wrap_align

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"KEKA_Asset_Export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

# === 📥 2. EXPORT EXCEL =======================================
@export_bp.route('/excel')
def export_excel():
    ids_param = request.args.get("ids")
    query = {}
    if ids_param:
        ids = [ObjectId(i) for i in ids_param.split(",") if i]
        query = {"_id": {"$in": ids}}

    assets = list(assets_collection.find(query))
    assets_by_type = {}
    for asset in assets:
        asset_type = str(asset.get("category") or "Unknown").strip()
        assets_by_type.setdefault(asset_type, []).append(asset)

    wb = Workbook()
    wb.remove(wb.active)

    def normalize_gst_key(key: str) -> str:
        if not key:
            return key
        if key.startswith("gst_"):
            return key.replace("(", "").replace(")", "").replace("%", "")
        return key

    def get_fields(asset_type: str):
        doc = asset_types_collection.find_one({"type_name": asset_type})
        if doc and "fields" in doc:
            fields = []
            for f in doc["fields"]:
                name = normalize_gst_key(f.get("name") or "")
                label = f.get("label") or ""
                if name.startswith("gst_"):
                    try:
                        rate = name.split("_")[1]
                        label = f"GST ({rate}%)"
                    except Exception:
                        label = "GST"
                fields.append({"label": label, "name": name, "type": f.get("type", "text")})
            return fields

        sample = assets_by_type.get(asset_type, [{}])[0]
        keys = [normalize_gst_key(k) for k in sample.keys() if k not in {"_id"}]

        fields = []
        for k in keys:
            if k.startswith("gst_"):
                try:
                    rate = k.split("_")[1]
                    label = f"GST ({rate}%)"
                except Exception:
                    label = "GST"
                fields.append({"label": label, "name": k, "type": "number"})
            else:
                fields.append({"label": k.replace("_", " ").title(), "name": k, "type": "text"})
        return fields

    def is_currency_field(field):
        name = (field.get("name") or "").lower()
        label = (field.get("label") or "").lower()
        type_ = (field.get("type") or "").lower()
        return name in {"amount", "total"} or name.startswith("gst_") or label.startswith("gst") or type_ == "number" and ("amount" in name or "total" in name or "gst" in name)

    def coerce_number(val):
        if val is None:
            return None
        if isinstance(val, (int, float)):
            return float(val)
        s = str(val).strip()
        if not s:
            return None
        s = s.replace("₹", "").replace(",", "").replace(" ", "")
        try:
            return float(s)
        except Exception:
            return None

    header_font = Font(bold=True, color="FFFFFF", name="Calibri")
    cell_font = Font(name="Calibri")
    fill = PatternFill(start_color="043251", end_color="043251", fill_type="solid")
    align_wrap = Alignment(wrap_text=True, vertical="top")

    for asset_type, asset_list in assets_by_type.items():
        fields = get_fields(asset_type)
        if not fields:
            continue

        sheet = wb.create_sheet(title=asset_type[:31])
        headers = [f["label"] for f in fields]
        keys = [f["name"] for f in fields]

        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = fill
            cell.alignment = align_wrap
            sheet.column_dimensions[cell.column_letter].width = 25

        for row_idx, asset in enumerate(asset_list, start=2):
            normalized_asset = {normalize_gst_key(k): v for k, v in asset.items()}
            for col_idx, field in enumerate(fields, start=1):
                key = field.get("name")
                label = field.get("label")
                dtype = (field.get("type") or "text").lower()

                raw_value = normalized_asset.get(key) or normalized_asset.get(label) or ""
                val = raw_value

                if dtype == "date" and val:
                    try:
                        dt = datetime.strptime(str(val), "%d-%m-%Y")
                        val = dt.strftime("%d-%m-%Y")
                    except Exception:
                        try:
                            dt = datetime.strptime(str(val), "%d/%m/%Y")
                            val = dt.strftime("%d-%m-%Y")
                        except Exception:
                            pass
                elif is_currency_field(field):
                    num = coerce_number(val)
                    if num is not None:
                        val = num

                sheet.cell(row=row_idx, column=col_idx, value=val).font = cell_font
                sheet.cell(row=row_idx, column=col_idx).alignment = align_wrap

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"Asset_Export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

#=== 📤 3. EXPORT MONGODB DATABASE ============================
@export_bp.route('/export_db')
def export_db():
    BACKUP_FOLDER = 'mongo_backups'
    DB_NAME = 'ams'
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    dump_path = os.path.join(BACKUP_FOLDER, f"{DB_NAME}_{timestamp}")

    try:
        os.makedirs(BACKUP_FOLDER, exist_ok=True)

        # Clean old backup
        if os.path.exists(dump_path):
            shutil.rmtree(dump_path)

        # Run mongodump
        MONGODUMP_PATH = r"C:\Users\Admin\Downloads\mongodb-tools\bin\mongodump.exe"
        subprocess.run([MONGODUMP_PATH, '--db', DB_NAME, '--out', BACKUP_FOLDER], check=True)


        flash('✅ MongoDB export completed successfully.', 'success')
    except Exception as e:
        flash(f'❌ Export failed: {e}', 'danger')

    return redirect(url_for('main.dashboard'))

# === 4. IMPORT EXCEL ======================================
@export_bp.route('/import_excel', methods=['POST'])
def import_excel():
    file = request.files.get('file')
    if not file:
        flash("❌ No file uploaded.", "danger")
        return redirect(url_for('main.dashboard'))

    try:
        wb = load_workbook(file, data_only=True)
    except Exception as e:
        flash(f"❌ Failed to read Excel file: {e}", "danger")
        return redirect(url_for('main.dashboard'))

    preview_data = []
    error_count = 0
    sheet_headers = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        # Skip empty sheet or header-only sheet
        if not rows or len(rows) < 2:
            continue

        headers = [str(h).strip() if h else "" for h in rows[0]]
        sheet_headers[sheet_name] = headers

        for idx, row in enumerate(rows[1:], start=2):
            # Skip fully empty rows
            if all((v is None or str(v).strip() == "") for v in row):
                continue

            row_data = OrderedDict()
            errors = {}
            suggestions = {}

            amount_val = None
            total_val  = None
            gst_rate = None
            gst_amount = None
            gst_header_name = None   # capture the exact GST header we used
            gst_seen_once = False    # enforce exactly one GST column

            # --- Build row data with pretty strings ---
            for col_index, header in enumerate(headers):
                value = row[col_index] if col_index < len(row) else ""
                h_low = header.lower()

                if h_low == "amount":
                    amount_val = safe_to_float(value, 0.0)
                    row_data[header] = format_inr(amount_val)

                elif h_low == "total":
                    total_val = safe_to_float(value, 0.0)
                    row_data[header] = format_inr(total_val)

                elif h_low.startswith("gst") and "(" in header and ")" in header:
                    # Expect exactly one GST column like "GST (18%)"
                    try:
                        rate_str = re.search(r"\((\d+)\s*%?\)", header).group(1)
                        this_rate = float(rate_str)
                        this_gst_amount = safe_to_float(value, 0.0)

                        if gst_seen_once:
                            # More than one GST column present → flag error
                            errors[header] = "Multiple GST columns not allowed"
                            # Keep whatever is in the cell as-is, but pretty if numeric
                            try:
                                row_data[header] = format_inr(this_gst_amount)
                            except Exception:
                                row_data[header] = str(value) if value is not None else ""
                        else:
                            gst_seen_once = True
                            gst_rate = this_rate
                            gst_amount = this_gst_amount
                            gst_header_name = header
                            row_data[header] = format_inr(this_gst_amount)
                    except Exception:
                        errors[header] = "Invalid GST header/value"
                        suggestions[header] = "₹0.00"
                        row_data[header] = str(value) if value is not None else ""

                else:
                    if value is None:
                        row_data[header] = ""
                    else:
                        row_data[header] = str(value).strip()

            # --- GST / Total validation (exactly one GST expected per row) ---
            if amount_val is not None and gst_rate is not None and gst_amount is not None:
                expected_gst = round(amount_val * gst_rate / 100, 2)

                if round(gst_amount, 2) != expected_gst and gst_header_name:
                    errors[gst_header_name] = f"GST mismatch ({int(gst_rate)}%)"
                    suggestions[gst_header_name] = f"Expected: {format_inr(expected_gst)}"

                if total_val is not None:
                    expected_total = round(amount_val + expected_gst, 2)
                    total_header = next((h for h in headers if h.lower() == "total"), None)
                    if total_header and round(total_val, 2) != expected_total:
                        errors[total_header] = "Total mismatch"
                        suggestions[total_header] = f"Expected: {format_inr(expected_total)}"

            # --- Dates check (DD-MM-YYYY). Future dates not allowed; no auto-fix ---
            for col_index, header in enumerate(headers):
                value = row[col_index] if col_index < len(row) else ""
                if "date" in header.lower() and value:
                    try:
                        if isinstance(value, datetime):
                            dt_obj = value
                        else:
                            dt_obj = datetime.strptime(str(value).strip(), "%d-%m-%Y")

                        if dt_obj > datetime.now():
                            errors[header] = "Future date not allowed"

                        # normalize preview as dd-mm-yyyy
                        row_data[header] = dt_obj.strftime("%d-%m-%Y")
                    except Exception:
                        errors[header] = "Invalid date format"

            if errors:
                error_count += 1

            preview_data.append({
                "sheet": sheet_name,
                "row": idx,
                "data": row_data,
                "errors": errors,
                "suggestions": suggestions
            })

    # --- Persist preview in Mongo, keep only preview_id in session ---
    try:
        preview_doc = {
            "created_at": datetime.utcnow(),
            "sheet_headers": sheet_headers,
            "preview_data": preview_data,
            "error_count": error_count,
        }
        result = import_previews_collection.insert_one(preview_doc)
        session["preview_id"] = str(result.inserted_id)
    except Exception as e:
        # Fallback if Mongo insert failed
        flash(f"❌ Failed to save preview: {e}", "danger")
        return redirect(url_for('main.dashboard'))

    # Render preview page (routes that follow read from Mongo using preview_id)
    return render_template(
        'import_preview.html',
        preview_data=preview_data,
        error_count=error_count,
        sheet_headers=sheet_headers
    )

@export_bp.route("/import_preview")
def import_preview():
    preview_id = session.get("preview_id")
    if not preview_id:
        flash("⚠️ No preview found. Please upload a file first.", "danger")
        return redirect(url_for("main.dashboard"))

    preview_doc = import_previews_collection.find_one({"_id": ObjectId(preview_id)})
    if not preview_doc:
        flash("⚠️ Preview expired or not found. Please re-upload the file.", "danger")
        return redirect(url_for("main.dashboard"))

    # Apply suggestions and format numbers for preview
    for row in preview_doc["preview_data"]:
        data = row["data"]
        suggestions = row.get("suggestions", {})
        for key, val in data.items():
            # Apply suggestion if it exists
            if key in suggestions:
                suggestion_val = suggestions[key]
                # Try to parse numeric suggestions
                match = re.search(r"₹([\d,]+(?:\.\d+)?)", str(suggestion_val))
                if match:
                    val = float(match.group(1).replace(",", ""))
                else:
                    try:
                        val = float(suggestion_val)
                    except:
                        val = suggestion_val
            # Format Amount, GST, Total nicely
            if key.lower() in ["amount", "total"] or re.match(r"gst", key.lower()):
                try:
                    val = format_inr(float(val)) if val not in [None, ""] else ""
                except:
                    val = val
            data[key] = val

    return render_template(
        "import_preview.html",
        preview_data=preview_doc["preview_data"],
        error_count=preview_doc["error_count"],
        sheet_headers=preview_doc["sheet_headers"],
    )


@export_bp.route('/confirm_import', methods=['POST'])
def confirm_import():
    preview_id = session.get("preview_id")
    if not preview_id:
        flash("⚠️ No preview found. Please upload a file first.", "danger")
        return redirect(url_for("main.dashboard"))

    preview_doc = import_previews_collection.find_one({"_id": ObjectId(preview_id)})
    if not preview_doc:
        flash("⚠️ Preview expired or not found. Please re-upload.", "danger")
        return redirect(url_for("main.dashboard"))

    preview_data = preview_doc["preview_data"]
    sheet_headers = preview_doc["sheet_headers"]
    assets = []

    def header_to_field_key(h):
        h_low = str(h).lower()
        if h_low.startswith("gst") and "(" in h and ")" in h:
            try:
                rate = re.search(r"\((\d+)\s*%?\)", h).group(1)
                return f"gst_{rate}"
            except Exception:
                return h
        if h_low.startswith("gst_"):
            return h_low
        return h

    for row in preview_data:
        sheet_name = row.get('sheet')
        headers = sheet_headers.get(sheet_name, list(row['data'].keys()))

        asset_type = asset_types_collection.find_one({"type_name": sheet_name})
        if not asset_type:
            new_fields = [{"label": h, "name": h.lower().replace(" ", "_")} for h in headers]
            asset_types_collection.insert_one({"type_name": sheet_name, "fields": new_fields})
            asset_type = asset_types_collection.find_one({"type_name": sheet_name})

        clean_data = OrderedDict()
        for h in headers:
            if h is None:
                continue  # skip None headers entirely
            schema_field = next((f for f in asset_type["fields"] if f["label"] == h), None)
            schema_key = schema_field["name"] if schema_field else h

            field_key = header_to_field_key(h)
            if field_key == h:
                field_key = schema_key

            v_pretty = row['data'].get(h, None)

            if not v_pretty or (isinstance(v_pretty, str) and v_pretty.strip() == ""):
                clean_data[field_key] = None
            else:
                if "date" in field_key.lower():
                    try:
                        dt_obj = datetime.strptime(str(v_pretty).strip(), "%d-%m-%Y")
                        clean_data[field_key] = dt_obj.strftime("%d-%m-%Y")
                    except Exception:
                        clean_data[field_key] = v_pretty
                else:
                    if field_key in ("amount", "total") or field_key.startswith("gst_"):
                        clean_data[field_key] = safe_to_float(v_pretty, 0.0)
                    else:
                        clean_data[field_key] = v_pretty

        # Remove any accidental None keys
        clean_data = {k: v for k, v in clean_data.items() if k is not None}

        clean_data = normalize_gst_keys(clean_data)
        clean_data["category"] = sheet_name
        assets.append(clean_data)

    if assets:
        assets_collection.insert_many(assets)
        flash(f"✅ Imported {len(assets)} assets.", "success")
    else:
        flash("⚠️ No assets to import.", "warning")

    import_previews_collection.delete_one({"_id": ObjectId(preview_id)})
    session.pop("preview_id", None)

    return redirect(url_for('main.dashboard'))

@export_bp.route('/download_errors', methods=['POST'])
def download_errors():
    preview_id = session.get("preview_id")
    if not preview_id:
        flash("⚠️ No preview found.", "danger")
        return redirect(url_for("main.dashboard"))

    preview_doc = import_previews_collection.find_one({"_id": ObjectId(preview_id)})
    if not preview_doc:
        flash("⚠️ Preview expired or not found.", "danger")
        return redirect(url_for("main.dashboard"))

    preview = preview_doc["preview_data"]

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['Sheet Name', 'Row Number', 'Field', 'Error'])

    for row in preview:
        for field in row.get('errors', []):   # safer with .get()
            writer.writerow([
                row['sheet'],
                row.get('row', '?'),
                field,
                'Validation error'
            ])

    response = make_response(output.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=error_report.csv'
    response.headers['Content-Type'] = 'text/csv'
    return response

@export_bp.route('/download_fixed', methods=['POST'])
def download_fixed():
    preview_id = session.get("preview_id")
    if not preview_id:
        flash("⚠️ No preview found.", "danger")
        return redirect(url_for("main.dashboard"))

    preview_doc = import_previews_collection.find_one({"_id": ObjectId(preview_id)})
    if not preview_doc:
        flash("⚠️ Preview expired or not found.", "danger")
        return redirect(url_for("main.dashboard"))

    preview_data = preview_doc["preview_data"]
    sheet_headers = preview_doc["sheet_headers"]

    output = BytesIO()
    wb = Workbook()
    wb.remove(wb.active)

    grouped = {}
    for row in preview_data:
        grouped.setdefault(row["sheet"], []).append(row)

    for sheet_name, rows in grouped.items():
        ws = wb.create_sheet(title=sheet_name)
        headers = sheet_headers.get(sheet_name, list(rows[0]["data"].keys()))
        ws.append(headers)

        for row in rows:
            fixed = []
            for key in headers:
                val = row["data"].get(key, "-")
                suggestion = row.get("suggestions", {}).get(key, None)

                # Apply suggestion if it exists
                if suggestion:
                    if not re.search(r"date", key.lower()):  # skip dates
                        match = re.search(r"₹([\d,]+(?:\.\d+)?)", suggestion)
                        if match:
                            val_str = match.group(1).replace(",", "")
                            try:
                                val = float(val_str)
                            except Exception:
                                val = val_str
                        else:
                            # For plain numeric suggestions without ₹
                            try:
                                val = float(suggestion)
                            except Exception:
                                val = suggestion
                    else:
                        val = suggestion  # date suggestions directly applied

                # Fallback for Amount, GST, Total if still empty
                if key.lower() in ["amount", "total"] or re.match(r"gst", key.lower()):
                    if val in [None, "", "-"]:
                        val = 0.0

                fixed.append(val)
            ws.append(fixed)

    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        download_name="fixed_import.xlsx",
        as_attachment=True,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# === 📥 5. IMPORT MONGODB DATABASE ===========================
@export_bp.route('/import_db')
def import_db():
    BACKUP_FOLDER = 'mongo_backups'
    DB_NAME = 'ams'
    restore_path = os.path.join(BACKUP_FOLDER, DB_NAME)

    try:
        folders = sorted(
            [f for f in os.listdir(BACKUP_FOLDER) if f.startswith(DB_NAME)],
            reverse=True
        )
        if not folders:
            flash('⚠️ No backup folders found.', 'warning')
            return redirect(url_for('dashboard'))

        latest = folders[0]
        restore_path = os.path.join(BACKUP_FOLDER, latest)

        MONGORESTORE_PATH = r"C:\Users\Admin\Downloads\mongodb-tools\bin\mongorestore.exe"
        subprocess.run([MONGORESTORE_PATH, '--drop', '--db', DB_NAME, restore_path], check=True)


        flash('✅ MongoDB import completed successfully.', 'success')
    except Exception as e:
        flash(f'❌ Import failed: {e}', 'danger')

    return redirect(url_for('main.dashboard'))

def run_weekly_backup():
    BACKUP_FOLDER = 'mongo_backups'
    DB_NAME = 'ams'
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    dump_path = os.path.join(BACKUP_FOLDER, f"{DB_NAME}_{timestamp}")

    try:
        os.makedirs(BACKUP_FOLDER, exist_ok=True)
        subprocess.run([
            r"C:\Users\Admin\Downloads\mongodb-tools\bin\mongodump.exe",
            '--db', DB_NAME, '--out', BACKUP_FOLDER
        ], check=True)
        print(f'✅ Weekly MongoDB backup completed at {timestamp}.')
    except Exception as e:
        print(f'❌ Weekly backup failed: {e}')

def start_backup_scheduler():
    schedule.every().week.do(run_weekly_backup)

    def run():
        while True:
            schedule.run_pending()
            time.sleep(60)

    threading.Thread(target=run, daemon=True).start()

def prepare_export_rows(assets):
    master_fields = get_master_fields()
    rows = []

    for asset in assets:
        row = []
        for field in master_fields:
            key = field["name"]
            label = field["label"]

            # Try canonical key first, fallback to label, else empty
            value = asset.get(key) or asset.get(label) or ""
            row.append(value)
        rows.append(row)

    return rows

@export_bp.route('/manual_backup')
def manual_backup():
    try:
        run_weekly_backup()
        flash("✅ Manual MongoDB backup completed successfully.", "success")
    except Exception as e:
        flash(f"❌ Manual backup failed: {e}", "danger")
    return redirect(url_for('main.dashboard'))

@export_bp.route('/bulk_delete', methods=['POST'])
def bulk_delete():
    ids_param = request.json.get("ids", [])
    if not ids_param:
        return jsonify({"error": "No IDs provided"}), 400

    from bson import ObjectId
    try:
        obj_ids = [ObjectId(i) for i in ids_param if i]
        result = assets_collection.delete_many({"_id": {"$in": obj_ids}})
        return jsonify({"message": f"✅ Deleted {result.deleted_count} assets"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500


start_backup_scheduler()
